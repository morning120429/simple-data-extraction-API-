from selenium import webdriver
from selenium.common.exceptions import *
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webelement import WebElement
from time import sleep
from openpyxl import Workbook
import json
import requests

def Find_Element(driver : webdriver.Chrome, by, value : str) -> WebElement:
    while True:
        try:
            element = driver.find_element(by, value)
            break
        except:
            pass
        sleep(0.1)
    return element

def Find_Elements(driver : webdriver.Chrome, by, value : str) -> list[WebElement]:
    while True:
        try:
            elements = driver.find_elements(by, value)
            if len(elements) > 0:
                break
        except:
            pass
        sleep(0.1)
    return elements

def Send_Keys(element : WebElement, content : str):
    element.clear()
    for i in content:
        element.send_keys(i)
        sleep(0.1)

def extractLink():
    resultOffset = 0
    resultRecordCount = 65000
    all_data = []

    while True:
        response = requests.get(f'https://geodata.md.gov/imap/rest/services/PlanningCadastre/MD_PropertyData/MapServer/0/query?where=1%3D1&outFields=SDATWEBADR&returnGeometry=false&resultOffset={resultOffset}&resultRecordCount={resultRecordCount}&outSR=4326&f=json')

        # Check the response status code.
        if response.status_code == 200:
            data = response.json()
            records = data.get("features", [])

            # Modify the key name 'SDATWEBADR' to 'link' in each record
            for record in records:
                attributes = record.get("attributes", {})
                if 'SDATWEBADR' in attributes:
                    attributes['link'] = attributes.pop('SDATWEBADR')

            # Add the records to the list.
            print(len(records))
            all_data.extend(records)
        else:
            # An error occurred.
            print("Error sending message: {}".format(response.status_code))

        if len(records) < resultRecordCount:
            break

        resultOffset += resultRecordCount

    transformed_data = []
    for item in all_data:
        attributes = item.get('attributes', {})
        transformed_data.append(attributes)

    with open('total_link.json', 'w') as file:
        json.dump(transformed_data, file, indent = 4)

    print(f'Extracted Total Link : {len(transformed_data)}')

def extractResult():
    driver = webdriver.Chrome()

    wb = Workbook()
    sheet = wb.active
    item = ['Account Identifier', 'Owner Name', 'Use', 'Premises Address', 'Primary Structure Built', 'Above Grade Living Area', 'Finished Basement Area', 'Property Land Area', 'Stores', 'Full/Half Bath', 'Value Information(Value)', 'Value Information(Date)', 'Transfer Information(Date)', 'Transfer Information(Price)', 'Link']
    for i in range(1, 16):
        sheet.cell(row = 1, column = i).value = item[i-1]

    with open('total_link.json', 'r') as file:
        data = json.load(file)

    start_row = 2
    output_json = []
    for i, element in enumerate(data, start = 1):
        print(f'{"-"*20} {i} {"-"*20}')
        driver.get(element['link'])

        table = Find_Element(driver, By.ID, 'detailSearch').find_element(By.TAG_NAME, 'tbody').find_elements(By.TAG_NAME, 'tr')

        account_identifier = table[3].find_element(By.TAG_NAME, 'span').text.replace('District - ', '').replace(' Account Number - ', '')
        print(f'Account Identifier : {account_identifier}')

        owner_name = table[5].find_elements(By.TAG_NAME, 'td')[1].text.split('\n')
        print(f'Owner Name : {", ".join(owner_name[0 : ])}')

        use = table[5].find_elements(By.TAG_NAME, 'td')[3].text.split('\n')
        print(f'Use : {use[0]}')

        premises_address = table[8].find_elements(By.TAG_NAME, 'td')[1].text
        print(f'Premises Address : {premises_address}')

        primary_structure_built = driver.find_element(By.XPATH, '//*[@id="detailSearch"]/tbody/tr[11]/td/table/tbody/tr[2]/td[1]').text
        print(f'Primary Structure Built : {primary_structure_built}')

        above_grade_living = driver.find_element(By.XPATH, '//*[@id="detailSearch"]/tbody/tr[11]/td/table/tbody/tr[2]/td[2]').text
        print(f'Above Grade Living Area : {above_grade_living}')

        finished_basement = driver.find_element(By.XPATH, '//*[@id="detailSearch"]/tbody/tr[11]/td/table/tbody/tr[2]/td[3]').text
        print(f'Finished Basement Area : {finished_basement}')

        property_land_area = driver.find_element(By.XPATH, '//*[@id="detailSearch"]/tbody/tr[11]/td/table/tbody/tr[2]/td[4]').text
        print(f'Property Land Area : {property_land_area}')

        stores = driver.find_element(By.XPATH, '//*[@id="detailSearch"]/tbody/tr[13]/td/table/tbody/tr[2]/td[1]').text
        print(f'Stores : {stores}')

        full_half_baths = driver.find_element(By.XPATH, '//*[@id="detailSearch"]/tbody/tr[13]/td/table/tbody/tr[2]/td[6]').text
        print(f'Full/Half Baths : {full_half_baths}')

        value_price = driver.find_element(By.XPATH, '//*[@id="detailSearch"]/tbody/tr[15]/td/table/tbody/tr[5]/td[5]').text
        print(f'Value Information(Value) : {value_price}')

        value_date = driver.find_element(By.XPATH, '//*[@id="detailSearch"]/tbody/tr[15]/td/table/tbody/tr[2]/td[5]').text.split('\n')[1]
        print(f'Value Information(Date)')

        transfer_date = driver.find_element(By.XPATH, '//*[@id="detailSearch"]/tbody/tr[17]/td/table/tbody/tr[1]/td[2]').find_element(By.TAG_NAME, 'span').text
        print(f'Transfer Date : {transfer_date}')

        transfer_price = driver.find_element(By.XPATH, '//*[@id="detailSearch"]/tbody/tr[17]/td/table/tbody/tr[1]/td[3]').find_element(By.TAG_NAME, 'span').text
        print(f'Transfer Price : {transfer_price}')

        sheet.cell(row = start_row, column = 1).value = account_identifier
        sheet.cell(row = start_row, column = 2).value = ', '.join(owner_name[0 : ])
        sheet.cell(row = start_row, column = 3).value = use[0]
        sheet.cell(row = start_row, column = 4).value = premises_address
        sheet.cell(row = start_row, column = 5).value = primary_structure_built
        sheet.cell(row = start_row, column = 6).value = above_grade_living
        sheet.cell(row = start_row, column = 7).value = finished_basement
        sheet.cell(row = start_row, column = 8).value = property_land_area
        sheet.cell(row = start_row, column = 9).value = stores
        sheet.cell(row = start_row, column = 10).value = full_half_baths
        sheet.cell(row = start_row, column = 11).value = value_price
        sheet.cell(row = start_row, column = 12).value = value_date
        sheet.cell(row = start_row, column = 13).value = transfer_date
        sheet.cell(row = start_row, column = 14).value = transfer_price
        sheet.cell(row = start_row, column = 15).value = element['link']

        output_json.append({
            'ID': account_identifier,
            'OwnerName': ', '.join(owner_name[0 : ]),
            'Use': use[0],
            'Address': premises_address,
            'YearBuilt': primary_structure_built,
            'LivingArea': above_grade_living,
            'FinishedBasement': finished_basement,
            'Stories': stores,
            'Full/Half Bath': full_half_baths,
            'LotSize': property_land_area,
            'AssessedValue': value_price,
            'AssessedDate': value_date,
            'LastSold': transfer_date,
            'LastSoldDate': transfer_price
            })
        
        with open('Result.json', 'w') as f:
            json.dump(output_json, f)

        wb.save('Result.xlsx')
        start_row += 1

if __name__ == '__main__':
    extractLink()
    extractResult()